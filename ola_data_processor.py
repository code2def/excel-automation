import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Define user mapping

# Full user ID to full name mapping
valid_users = {
    "akumar3": "Ajay",
    "asingh55": "Akhilendra",
    "agiri1": "Aman",
    "aar": "Amar A R",
    "akadyan": "Amit Kadyan",
    "akumar72": "Amit Kumar",
    "araina": "Amit Raina",
    "aagarwal1": "Amrish",
    "abharti": "Ankur",
    "adubey3": "Anu Dubey",
    "agovindan": "Arul",
    "amathur2": "Ashish Mathur",
    "dahuja": "Daksh Ahuja",
    "dharit": "Deepak Harit",
    "dmam": "Deepak Mam",
    "dprakash": "Dharma Prakash",
    "dpanwar": "Dharmesh Panwar",
    "hkumar2": "Hemant Kumar",
    "bbindhu": "B. Hima Bindhu",
    "ksharma5": "Kapil Sharma",
    "kbabu": "Karthick Babu",
    "ksinghpatel": "Kunj Bihari",
    "mranganathan": "Magesh Ranganathan",
    "mgundev": "Manish Gundev",
    "mfaraz": "Mohammed Faraz",
    "npadhi": "Nalini Padhi",
    "nnanjappa": "Nijagunashivayogi",
    "psrihari": "Prakasam Srihari",
    "psrivastav": "Pranshu Srivastav",
    "rhalsi": "Rekha halsi",
    "rjain6": "Rohit Jain",
    "sjain3": "Salil Jain",
    "skumar25": "Sandeep Kumar",
    "sjoshi3": "Sanjana Joshi",
    "syadav1": "Saurabh Yadav",
    "sbhargava": "Shailesh Bhargava",
    "skumar51": "Shishir Kumar",
    "sjain16": "Siddharth Jain",
    "spatnam": "Sreekanth Patnam",
    "sshukla1": "Sunil Shukla",
    "tgupta1": "Tushar Gupta",
    "mmohammed1": "Umar Mohammed",
    "vchoudhary": "Varun Choudhary",
    "vdixit": "Vipul Dixit",
    "aali2": "Adnan Ali",
    "cpradhan": "Chandan Kumar Pradhan",
    "dnai": "Deepak Nai",
    "dsonker": "Dilip Sonker",
    "gch": "Gowri CH",
    "ganthwal": "Gunjan Anthwal",
    "hgambhir": "Himanshu Gambhir",
    "kyadav": "Khushboo Yadav",
    "mashraf": "Mohammed Aamir Ashraf",
    "rkundu": "Rohan Kundu",
    "syadav14": "Sangeeta Yadav",
    "smalhotra5": "Satish Malhotra",
    "mmannan": "Shahid Mohammad",
    "schakkalackalponnapp": "Shybin",
    "snegi": "Swati Negi",
    "tjangra": "Tarun Jangra",
    "asingh21": "Awanish Singh",
    "mmittal": "Mohit Mittal",
    "skumar19": "Satish Kumar",
    "tpasha": "Tabrez Pasha",
    "ybhatti": "Yameen Bhatti",
    "vkumar": "Vipin Kumar",
    "vkaul11": "Varesh Kaul",
    "sarikapudi": "Sudheer Arikapudi",
    "ckk": "Chaithra",
    "asingh8": "Alok Kumar Singh",
    "amalhotra10": "Arun Malhotra",
    "cjayaraman": "Chandrasekaran Jayaraman",
    "amishra2": "Atul Mishra"
}

def process_excel(file):
    df = pd.read_excel(file, engine='pyxlsb')

    # Apply filters
    filtered_df = df[
        (df["QUEUE_CODE"] == "BDWCNFG") &
        (df["D_IN_OUT_OLA"] == "OUT OF OLA") &
        (df["USER_ID_COMPLETION"].isin(valid_users.keys()))
    ].copy()
    
    # Add Failure category & Failure Reasons
    if "Failure category" not in filtered_df.columns:
        filtered_df["Failure category"] = ""
    if "Failure Reasons" not in filtered_df.columns:
        filtered_df["Failure Reasons"] = ""
    
    # Update logic for Failure category and Failure Reasons
    filtered_df.loc[filtered_df["DELAY_DIARY"].isna(), "Failure category"] = "Genuine Fault / Prioritization Error"
    filtered_df.loc[filtered_df["DELAY_DIARY"].isna(), "Failure Reasons"] = filtered_df["USER_ID_COMPLETION"].map(
        lambda x: f"Missed to close on time by {valid_users.get(x, x)}"
    )
    
    # Define the correct column order
    correct_order = [
        "QUEUE_CODE", "TASK_CREATE", "TASK_CLOSED", "NEW_CONTRACT_NO", "ORDER_TYPE", "ORDER_TYPE_GROUP_CALC",
        "D_ORDER_TYPE_GROUP", "COUNTRY", "WORK_ITEM_ID_CALC", "REPORTING_WEEK", "REPORTING_MONTH",
        "PRODUCT_OFFERING", "D_OLA_TARGET", "LEAD_TIME_OVERALL", "D_IN_OUT_OLA", "REJECTION_DURATION_OVERALL",
        "CUSTOMER_ON_HOLD_DURITON_CALC", "TASK_DELAY_DURATION_OVERALL", "IN_OUT_OLA", "TASK_DELAY_REASON",
        "CUSTOMER_ON_HOLD_REASON", "NIGO_REASON_1", "NIGO_REASON_2", "SERVICE_GROUP_CALC", "TOP_250",
        "OPERATOR_NAME", "USER_ID_COMPLETION", "NO_OF_CIRCUITS", "INSTALLATION_COUNTRY", "CUSTOMER_NAME",
        "CUSTOMER_TIER", "NIGO_REASON", "INSTALLATION_COUNTRY_A", "INSTALLATION_COUNTRY_B", "A_COUNTRY",
        "B_COUNTRY", "SYS_FLAG", "D_CONNECTION_TYPE", "D_CONNECTION_TYPE_2", "CONTRACT_PRIORITY_ALL",
        "D_BANDWIDTH", "SD_WAN_SITE_TYPE", "FAST_TRACK", "FASTRACK_SIEBEL", "SUBTYPE", "DELAY_DIARY",
        "DELAY_REASON", "SIEBEL_PROJECT_ID", "SIEBEL_PROJECT_MANAGER", "OHS_PROJECT_ID", "OHS_PROJECT_NAME",
        "Location", "Sub Team", "Failure category", "Failure Reasons", "Comments", "Team Name", "Column2"
    ]
    
    # Reorder columns if all are present
    filtered_df = filtered_df.reindex(columns=correct_order, fill_value="")
    
    return filtered_df

def format_excel(df):
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    # Write the DataFrame to the worksheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Apply formatting
    font = Font(name='Arial', size=9)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    alignment = Alignment(horizontal='left', vertical='center')

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = font
            cell.border = border
            cell.alignment = alignment

    # Format date columns as m/d/yyyy
    date_columns = ["TASK_CREATE", "TASK_CLOSED", "REPORTING_MONTH"]
    for col in date_columns:
        if col in df.columns:
            col_idx = df.columns.get_loc(col) + 1  # +1 because Excel columns start at 1
            for cell in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for c in cell:
                    c.number_format = 'm/d/yyyy'

    return wb

# Streamlit UI
st.title("OLA Data Processor")
uploaded_files = st.file_uploader("Upload Excel Files", type=["xlsb"], accept_multiple_files=True)

if uploaded_files:
    st.write("Processing files...")

    # Initialize an empty list to store filtered DataFrames
    filtered_dfs = []

    # Process each uploaded file
    for uploaded_file in uploaded_files:
        filtered_df = process_excel(uploaded_file)
        filtered_dfs.append(filtered_df)

    # Combine all filtered DataFrames into one
    if filtered_dfs:
        consolidated_df = pd.concat(filtered_dfs, ignore_index=True)

        st.write("### Consolidated Filtered Data Preview:")
        st.dataframe(consolidated_df)

        # Format and download the Excel file
        wb = format_excel(consolidated_df)
        output_filename = "consolidated_filtered_data.xlsx"
        wb.save(output_filename)

        with open(output_filename, "rb") as file:
            st.download_button(
                label="Download Consolidated Excel File",
                data=file,
                file_name=output_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        st.warning("No valid data found in the uploaded files.")
